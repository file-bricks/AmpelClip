import sys
from pathlib import Path
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))
from Ampel6 import read_first_column_values


def test_excel_nan_cells_excluded_from_word_list(tmp_path):
    """Leere Excel-Zellen dürfen nicht als 'nan'-String in die Wortliste gelangen."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Müller"
    ws["A2"] = None      # leere Zelle → NaN beim Einlesen
    ws["A3"] = "Schmidt"
    wb.save(tmp_path / "sensibel.xlsx")
    content = read_first_column_values(tmp_path / "sensibel.xlsx")
    assert "nan" not in content, "Leerzellen dürfen nicht als 'nan' erscheinen"
    assert "Müller" in content
    assert "Schmidt" in content
    assert len(content) == 2


def test_excel_without_dropna_would_produce_nan(tmp_path):
    """Bestätigt, dass das alte Verhalten (ohne dropna) 'nan' erzeugt hätte."""
    import pandas as pd
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Eintrag"
    ws["A2"] = None      # leere Zelle in der MITTE — pandas liest sie als NaN
    ws["A3"] = "Ende"    # Folgedaten zwingen pandas, die Leerzeile einzuschließen
    wb.save(tmp_path / "test.xlsx")
    df = pd.read_excel(tmp_path / "test.xlsx", header=None)
    broken = df.iloc[:, 0].astype(str).tolist()   # altes Verhalten ohne dropna
    assert "nan" in broken, "Ohne dropna() enthält die Liste 'nan' — das ist der Bug"
    # read_first_column_values filtert die NaN-Zelle korrekt heraus:
    content = read_first_column_values(tmp_path / "test.xlsx")
    assert "nan" not in content
    assert len(content) == 2


def test_rot_status_text_has_correct_capitalization():
    """Stellt sicher, dass der ROT-Statustext korrekte Großschreibung hat."""
    source = (Path(__file__).parent.parent / "Ampel6.py").read_text(encoding="utf-8")
    assert "Keine Änderung am Clipboard" in source, (
        "ROT-Statustext muss 'Keine Änderung' enthalten (Ä groß)"
    )
    assert "Keine änderung am Clipboard" not in source, (
        "Kleingeschriebenes 'änderung' gefunden — Tippfehler nicht behoben"
    )
